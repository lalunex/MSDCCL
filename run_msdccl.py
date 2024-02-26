import os

import recbole.utils.logger
import torch

from config.pre_train_model import pre_train_model_dict
import openpyxl
import yaml
from logging import getLogger
from recbole.utils import init_logger, init_seed, get_trainer
from recbole.config import Config
from recbole.data import create_dataset, data_preparation
from model.msdccl import MSDCCL

from Triainer.MsdcclTrainer import MsdcclTrainer
from utils.utils import get_model, is_msdccl_model
from utils.parser import parse_args


def main(
        model=None, dataset=None, config_dict=None, config_file_list=None, verbose=True, saved=True
):
    embedding_size = config_dict['embedding_size']
    note = 'hyper-parameter analysis, embedding size:[%d]' % embedding_size
    print('!!!Note: ' + note)

    config = Config(
        model=MSDCCL if is_msdccl_model(model) else model,
        dataset=dataset,
        config_file_list=config_file_list,
        config_dict=config_dict
    )
    init_seed(config['seed'], config['reproducibility'])

    # logger initialization
    init_logger(config)
    logger = getLogger()

    logger.info(config)

    # dataset filtering
    dataset = create_dataset(config)
    logger.info(dataset)

    # dataset splitting
    train_data, valid_data, test_data = data_preparation(config, dataset)
    # config['total_train_examples'] = train_data.sample_size
    config['total_train_examples'] = train_data.pr_end

    # model loading and initialization
    model_class = get_model(config_dict['model_name'])
    model = model_class(config, train_data.dataset).to(config['device'])
    logger.info(model)

    # trainer loading and initialization
    trainer = MsdcclTrainer(config, model)

    best_valid_score, best_valid_result = trainer.fit(
        train_data, valid_data, verbose=verbose, saved=saved, show_progress=config['show_progress']
    )

    # model evaluation
    test_result = trainer.evaluate(test_data)
    # test_result = trainer.evaluate(test_data, model_file='saved/MSDCCL-Dec-03-2023_21-57-02.pth')

    logger.info('best valid result: {}'.format(best_valid_result))
    logger.info('test result: {}'.format(test_result))

    if 'MSDCCL' in config_dict['model_name']:
        model_name_full = config_dict['model_name'] + '_' + config_dict['sub_model']
    else:
        model_name_full = config_dict['model_name']
    rst_dic = {
        'model': model_name_full,
        'best_valid_score': best_valid_score,
        'valid_score_bigger': config['valid_metric_bigger'],
        'best_valid_result': best_valid_result,
        'test_result': test_result
    }

    record_model_result(dataset, model_name_full, config_dict, rst_dic, note, config_file_list)
    return rst_dic


def get_parameter_dict(dic):
    parameter_dict = {
        'loss_type': 'BPR',
        "train_neg_sample_args": {
            'distribution': 'uniform',
            'sample_num': 1
        },

        # bert4rec 最大序列长度200 for ml-1m，50 for others
        "MAX_ITEM_LIST_LENGTH": 200 if dic['dataset'] == 'ml-1m' else 50,
        "pre_train_model_dict": pre_train_model_dict,

        # 下面是添加的caser代码来读取短期信息，初始化参数
        "user_short_nh": 8,
        "user_short_nv": 4,
        'transformer_encoder_heads': 2,
        'transformer_encoder_layers': 2,
        'transformer_encoder_dim_feedforward': 2048,
        'transformer_encoder_layer_norm_eps': 1e-12
    }
    parameter_dict.update(dic)

    if dic['dataset'] == 'yelp':
        # 下面我们从2018年开始提取数据
        parameter_dict['val_interval'] = {'timestamp': '[1546272000, inf]'}
    if dic['model_name'] == 'BERT4Rec' or (is_msdccl_model(dic['model_name']) and dic['sub_model'] == 'BERT4Rec'):
        bert_dict = {
            "embedding_size": dic['embedding_size'],
            "wandb_project": "recbole",
            "require_pow": False,
            "n_layers": 2,
            "n_heads": 2,
            "hidden_size": dic['embedding_size'],
            "inner_size": dic['embedding_size'],
            "hidden_dropout_prob": 0.5,
            "attn_dropout_prob": 0.5,
            "hidden_act": 'gelu',
            "layer_norm_eps": 1e-12,
            "mask_ratio": 0.2
        }
        parameter_dict.update(bert_dict)
    elif dic['model_name'] == 'GRU4Rec' or (is_msdccl_model(dic['model_name']) and dic['sub_model'] == 'GRU4Rec'):
        gru_dict = {
            "embedding_size": dic['embedding_size'],
            "hidden_size": dic['embedding_size'],
            "num_layers": 1,
            "dropout_prob": 0.3
        }
        parameter_dict.update(gru_dict)

    elif dic['model_name'] == 'SASRec' or (is_msdccl_model(dic['model_name']) and dic['sub_model'] == 'SASRec'):
        sas_dict = {
            "n_layers": 2,
            "n_heads": 2,
            "hidden_size": dic['embedding_size'],
            "inner_size": dic['embedding_size'],
            "hidden_dropout_prob": 0.5,
            "attn_dropout_prob": 0.5,
            "hidden_act": 'gelu',
            "layer_norm_eps": 1e-12,
            "initializer_range": 0.02,
            "loss_type": 'BPR',
            'neg_sampling': {
                'uniform': 5
            }
        }
        parameter_dict.update(sas_dict)

    elif dic['model_name'] == 'Caser' or (is_msdccl_model(dic['model_name']) and dic['sub_model'] == 'Caser'):
        caser_dict = {
            "embedding_size": dic['embedding_size'],
            "hidden_size": dic['embedding_size'],
            "dropout_prob": 0.4,
            "reg_weight": 1e-4,
            "weight_decay": 0.0,
            "nv": 4,
            "nh": 8,
            # "loss_type": 'CE',
            "MAX_ITEM_LIST_LENGTH": 5,
        }
        parameter_dict.update(caser_dict)

    elif dic['model_name'] == 'NARM' or (is_msdccl_model(dic['model_name']) and dic['sub_model'] == 'NARM'):
        narm_dict = {
            "embedding_size": dic['embedding_size'],
            "hidden_size": dic['embedding_size'],
            "n_layers": 1,
            "dropout_probs": [0.25, 0.5],
            # "loss_type": 'CE'
        }
        parameter_dict.update(narm_dict)

    elif dic['model_name'] == 'STAMP' or (is_msdccl_model(dic['model_name']) and dic['sub_model'] == 'STAMP'):
        stamp_dict = {
            "embedding_size": dic['embedding_size'],
            "hidden_size": dic['embedding_size'],
            # "loss_type": 'CE'
        }
        parameter_dict.update(stamp_dict)
    return parameter_dict


def record_model_result(dataset, model_name_full, config_dict, rst_dic, note, config_file_list):
    if not os.path.exists(output_file_name):
        wb = openpyxl.Workbook()
        sheet = wb.create_sheet()
    else:
        wb = openpyxl.load_workbook(output_file_name)
        sheet = wb['Sheet1']

    config = _load_config_files(config_file_list)
    eval_model = config['eval_args']['mode']
    dataset = dataset.dataset_name
    dataset += ('/' + eval_model) if eval_model is not None else ''
    run_rst = [dataset, model_name_full]
    valid_rst = list(rst_dic['best_valid_result'].values())
    run_rst.extend(valid_rst)
    run_rst.append('\t')
    test_rst = list(rst_dic['test_result'].values())
    run_rst.extend(test_rst)
    run_rst.append('\t')
    run_rst.append(str(config_dict))
    run_rst.append(note)
    sheet.append(run_rst)

    wb.save(output_file_name)


def _load_config_files(file_list):
    loader = yaml.FullLoader
    file_config_dict = dict()
    if file_list:
        for file in file_list:
            with open(file, 'r', encoding='utf-8') as f:
                file_config_dict.update(yaml.load(f.read(), Loader=loader))
    return file_config_dict


output_file_name = 'model_experiment_output-yelp.xlsx'

if __name__ == '__main__':
    config_file_list = ['config/text.yml']

    args = parse_args()
    dic = vars(args)
    parameter_dict = get_parameter_dict(dic)

    # parameter_dict['seed'] = 2021
    # parameter_dict['gpu_id'] = 2
    # parameter_dict['reweight_loss_lambda'] = 1

    main(model=dic['model_name'],
         dataset=dic['dataset'],
         config_dict=parameter_dict,
         config_file_list=config_file_list,
         verbose=dic['verbose'],
         saved=True)

    # seed_list = [2014,2015,2016,2017,2018]
    # seed_list = [2019, 2020, 2021, 2022, 2023]
    # # seed_list = [6,7,8,9,10]
    # reweight_loss_lambda = [1.8, 2]
    # # sequence_last_m = [4,5,6]
    # for seed in seed_list:
    #     for loss_lambda in reweight_loss_lambda:
    #         parameter_dict['seed'] = seed
    #         parameter_dict['reweight_loss_lambda'] = loss_lambda
    #         parameter_dict['gpu_id'] = 0
    #
    #         main(model=dic['model_name'],
    #              dataset=dic['dataset'],
    #              config_dict=parameter_dict,
    #              config_file_list=config_file_list,
    #              verbose=dic['verbose'],
    #              saved=True)

    # reweight_loss_lambda = [.2,.4,.6,.8,1,1.2,1.4,1.6,1.8,2]
    # reweight_loss_lambda = [1.2,1.4,1.6]
    # # for last_m in sequence_last_m:

    # # sequence_last_m = [4,5,6]
    # for loss_lambda in reweight_loss_lambda:
    #     parameter_dict['seed'] = 2021
    #     parameter_dict['gpu_id'] = 3
    #     parameter_dict['reweight_loss_lambda'] = loss_lambda

    #     main(model=dic['model_name'],
    #         dataset=dic['dataset'],
    #         config_dict=parameter_dict,
    #         config_file_list=config_file_list,
    #         verbose=dic['verbose'],
    #         saved=True)
    # reweight_loss_lambda = [.4,.6,.8,1]
    # for last_m in sequence_last_m:
    # sequence_last_m = [1,2,3,4,5,6]
    # seed_list = [2014,2015,2016,2017,2018,2019,2020,2021,2022,2023]
    # seed_list = [2022, 2023]
    # # seed_list = [2019,2020,2021,2022,2023]
    # # # reweight_loss_lambda = [.2,.4,.6,.8,1]
    # # # reweight_loss_lambda = [1.2,1.4,1.6,1.8,2]
    # # # reweight_loss_lambda = [.6,.8,1,1.2,1.4,1.6,1.8,2]
    # for seed in seed_list:
    #     parameter_dict['seed'] = seed
    #     parameter_dict['gpu_id'] = 0
    #     # parameter_dict['sequence_last_m'] = 6
    #
    #     main(model=dic['model_name'],
    #          dataset=dic['dataset'],
    #          config_dict=parameter_dict,
    #          config_file_list=config_file_list,
    #          verbose=dic['verbose'],
    #          saved=True)

    # seed_list = [2015]
    # sequence_last_m = [5,6]
    # for seed in seed_list:
    #   for last_m in sequence_last_m:
    #         parameter_dict['seed'] = seed
    #         parameter_dict['sequence_last_m'] = last_m
    #         parameter_dict['gpu_id'] = 3

    #         main(model=dic['model_name'],
    #             dataset=dic['dataset'],
    #             config_dict=parameter_dict,
    #             config_file_list=config_file_list,
    #             verbose=dic['verbose'],
    #             saved=True)

    # seed_list = [2016,2017]
    # sequence_last_m = [1,2,3,4,5,6]
    # for seed in seed_list:
    #   for last_m in sequence_last_m:
    #         parameter_dict['seed'] = seed
    #         parameter_dict['sequence_last_m'] = last_m
    #         parameter_dict['gpu_id'] = 3

    #         main(model=dic['model_name'],
    #             dataset=dic['dataset'],
    #             config_dict=parameter_dict,
    #             config_file_list=config_file_list,
    #             verbose=dic['verbose'],
    #             saved=True)

    # weight_decay choice [0,1e-3,1e-4]
    # our_att_drop_out choice [0,.1,.2,.3,.4,.5,.6,.7,.8,.9]
    # our_ae_drop_out choice [0,.1,.2,.3,.4,.5,.6,.7,.8,.9]
    # sequence_last_m choice [1,2,3,4,5,6,7,8,9,10]
    # reweight_loss_lambda choice [.2,.4,.6,.8,1,1.2,1.4,1.6,1.8,2]
    # record_model_result(output_string)
